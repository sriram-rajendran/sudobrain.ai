"""Gmail attachment download + text/OCR extraction."""

import base64
import logging
import shutil
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

logger = logging.getLogger("sudobrain.gmail.attachments")

TEXT_EXTRACTABLE = {"pdf", "docx", "txt", "md", "csv", "xlsx"}
IMAGE_TYPES = {"png", "jpg", "jpeg", "gif", "webp", "svg"}


def download_attachment(service, message_id: str, attachment_id: str) -> bytes:
    """Download raw attachment bytes via Gmail API."""
    resp = service.users().messages().attachments().get(
        userId="me", messageId=message_id, id=attachment_id
    ).execute()
    data = resp.get("data", "")
    return base64.urlsafe_b64decode(data) if data else b""


def extract_text(filename: str, kind: str, raw: bytes) -> str:
    """Extract plain text from attachment bytes. Returns '' on failure."""
    text, _, _ = extract_content(filename, kind, raw)
    return text


def extract_content(filename: str, kind: str, raw: bytes) -> tuple[str, str, str]:
    """Extract text and return (text, status, reason)."""
    if not raw:
        return "", "empty", "attachment_bytes_empty"
    k = kind.lower().lstrip(".")
    try:
        if k == "pdf":
            import fitz  # PyMuPDF
            doc = fitz.open(stream=raw, filetype="pdf")
            try:
                parts = [page.get_text() for page in doc]
            finally:
                doc.close()
            text = "\n\n".join(parts).strip()
            return _content_result(text, "pdf_text")
        if k == "docx":
            from docx import Document
            doc = Document(BytesIO(raw))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return _content_result(text, "docx_text")
        if k == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets:
                parts.append(f"## Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        parts.append("\t".join(cells))
            wb.close()
            return _content_result("\n".join(parts), "xlsx_text")
        if k in {"txt", "md", "csv"}:
            try:
                return _content_result(raw.decode("utf-8", errors="replace"), f"{k}_text")
            except Exception:
                return _content_result(raw.decode("latin-1", errors="replace"), f"{k}_latin1_text")
        if k in IMAGE_TYPES:
            return extract_image_text(filename, k, raw)
    except Exception as e:
        logger.warning("extract_text failed for %s (%s): %s", filename, k, e)
        return "", "failed", str(e)[:240]
    return "", "unsupported", f"unsupported_type:{k}"


def extract_image_text(filename: str, kind: str, raw: bytes) -> tuple[str, str, str]:
    """OCR an image attachment when a local OCR backend is available."""
    if kind == "svg":
        try:
            text = raw.decode("utf-8", errors="replace")
        except Exception:
            text = ""
        return _content_result(text, "svg_text")

    try:
        import pytesseract
        from PIL import Image

        image = Image.open(BytesIO(raw))
        text = pytesseract.image_to_string(image).strip()
        return _content_result(text, "pytesseract")
    except ModuleNotFoundError:
        pass
    except Exception as e:
        logger.debug("pytesseract OCR failed for %s: %s", filename, e)

    tesseract = shutil.which("tesseract")
    if not tesseract:
        return "", "ocr_unavailable", "tesseract_not_installed"

    suffix = f".{kind}"
    with tempfile.TemporaryDirectory(prefix="sudobrain_gmail_ocr_") as tmp:
        image_path = Path(tmp) / f"attachment{suffix}"
        image_path.write_bytes(raw)
        try:
            result = subprocess.run(
                [tesseract, str(image_path), "stdout", "--psm", "6"],
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
        except Exception as e:
            return "", "failed", f"tesseract_failed:{str(e)[:200]}"
        if result.returncode != 0:
            return "", "failed", f"tesseract_exit_{result.returncode}:{result.stderr[:160]}"
        return _content_result(result.stdout.strip(), "tesseract")


def _content_result(text: str, reason: str) -> tuple[str, str, str]:
    cleaned = (text or "").strip()
    if cleaned:
        return cleaned, "extracted", reason
    return "", "no_text", reason


def init_attachment_table():
    from backend.storage.database import get_connection
    conn = get_connection()
    try:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS gmail_attachments (
                id SERIAL PRIMARY KEY,
                message_id TEXT NOT NULL,
                filename TEXT,
                file_type TEXT,
                content_kind TEXT DEFAULT 'document',
                extraction_status TEXT DEFAULT 'pending',
                extraction_reason TEXT,
                size_kb REAL,
                extracted_text TEXT,
                char_count INTEGER DEFAULT 0,
                created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(message_id, filename)
            );
            ALTER TABLE gmail_attachments ADD COLUMN IF NOT EXISTS content_kind TEXT DEFAULT 'document';
            ALTER TABLE gmail_attachments ADD COLUMN IF NOT EXISTS extraction_status TEXT DEFAULT 'pending';
            ALTER TABLE gmail_attachments ADD COLUMN IF NOT EXISTS extraction_reason TEXT;
            CREATE INDEX IF NOT EXISTS idx_gmail_attach_msg ON gmail_attachments(message_id);
        """)
        conn.commit()
    finally:
        conn.close()


def store_attachment(message_id: str, filename: str, file_type: str,
                     size_kb: float, text: str, extraction_status: str,
                     extraction_reason: str):
    from backend.storage.database import get_connection
    init_attachment_table()
    conn = get_connection()
    try:
        conn.execute("""
            INSERT INTO gmail_attachments
            (message_id, filename, file_type, content_kind, extraction_status, extraction_reason,
             size_kb, extracted_text, char_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (message_id, filename) DO UPDATE SET
                content_kind = EXCLUDED.content_kind,
                extraction_status = EXCLUDED.extraction_status,
                extraction_reason = EXCLUDED.extraction_reason,
                size_kb = EXCLUDED.size_kb,
                extracted_text = EXCLUDED.extracted_text,
                char_count = EXCLUDED.char_count
        """, (
            message_id,
            filename,
            file_type,
            "image" if file_type.lower().lstrip(".") in IMAGE_TYPES else "document",
            extraction_status,
            extraction_reason[:240] if extraction_reason else None,
            size_kb,
            text,
            len(text),
        ))
        conn.commit()
    finally:
        conn.close()


def process_message_attachments(service, message_id: str,
                                attachments: list[dict]) -> int:
    """Download + extract + store text from a message's useful attachments.

    Returns count of attachments stored with non-empty text.
    """
    if not attachments:
        return 0
    init_attachment_table()
    stored = 0
    for a in attachments:
        kind = (a.get("type") or "").lower()
        if kind not in TEXT_EXTRACTABLE and kind not in IMAGE_TYPES:
            continue
        aid = a.get("attachment_id")
        if not aid:
            continue
        try:
            raw = download_attachment(service, message_id, aid)
            text, status, reason = extract_content(a.get("name", ""), kind, raw)
            store_attachment(
                message_id,
                a.get("name", ""),
                kind,
                float(a.get("size_kb", 0)),
                text,
                status,
                reason,
            )
            if text:
                stored += 1
        except Exception as e:
            logger.warning("process attachment %s/%s failed: %s",
                           message_id, a.get("name"), e)
    return stored
